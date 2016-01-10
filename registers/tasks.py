from __future__ import absolute_import
from django.core.exceptions import ObjectDoesNotExist
from openpyxl import load_workbook
import os
from unidecode import unidecode
from xlrd import open_workbook

from harvest.utils_pdq import csv_data
from tracking.models import Computer, Mobile
from .models import LocalPropertyRegister, LprValidation


def validate_lpr(lpr_id):
    from tracking.utils import logger_setup  # Avoid circular import.
    logger = logger_setup('task_validate_lpr')
    """Async task that accepts a LocalPropertyRegister object, and then
    iterates through the uploaded spreadsheet to validate data and record any
    failures in the Incredibus database.
    """
    try:
        lpr = LocalPropertyRegister.objects.get(pk=lpr_id)
    except ObjectDoesNotExist:
        logger.error('LocalPropertyRegister object not found ({})'.format(lpr_id))
        return False

    # Read source data into an iterable, prior to validation.
    # Data standard:
    #  - Uploaded file must be XLS, XLSX or CSV (validated by upload form).
    #  - [Excel] Data to be validated is on the first worksheet.
    #  - The first row of data is always a header row (skipped).
    #  - Column 1 contains device serial numbers, in text format.
    #  - Column 2 contains the assigned cost centre number, in integer format.
    logger.info('Handling {}'.format(lpr.uploaded_file.name))
    # Use the csv standard lib for CSV files.
    if os.path.splitext(lpr.uploaded_file.name)[1] == '.csv':
        rows = sum(1 for line in open(lpr.uploaded_file.path))
        data = csv_data(lpr.uploaded_file.path)
        data.next()  # Skip header row.
        data = [i for i in data]
    # Use the openpyxl lib for XLSX files
    elif os.path.splitext(lpr.uploaded_file.name)[1] == '.xlsx':
        workbook = load_workbook(filename=lpr.uploaded_file.path, read_only=True)
        sheet = workbook.active  # Sheet index 0.
        data = []
        for row in sheet.rows:
            data.append([cell.value for cell in row])
        data.pop(0)  # Pop out the header row.
    # Use the xlrd lib for XLS files.
    else:
        workbook = open_workbook(lpr.uploaded_file.path)
        sheet = workbook.sheet_by_index(0)
        rows = sheet.nrows
        data = []
        for i in range(1, rows):  # Skip header row.
            data.append([sheet.cell(i, 0).value, sheet.cell(i, 1).value])

    # Row validation rules:
    #  * Column 1 is a serial number.
    #  * Column 1 cannot be empty/blank.
    #  * Column 1 has a maximum length of 128 characters.
    #  * Column 2 is a cost centre number.
    #  * Column 2 cannot be empty/blank.
    #  * Column 2 must be interpreted as an integer value.

    for row_num, row in enumerate(data, 2):
        # Note that we start row_num at 2 to make sense for validation messages.
        failures = []
        # Serial number column
        if not row[0]:
            failures.append('Serial number is blank for row {}'.format(row_num))
            logger.warn('Serial number is blank for row {}'.format(row_num))
        else:
            # Account for text-encodings (Latin1), trailing spaces, etc.
            serial = row[0].replace(u'\xa0', u' ')
            serial = unidecode(serial).strip().upper()
            if len(serial) > 128:
                failures.append('Serial number is greater than 128 characters')
                logger.warn('Serial number is >128 chars for row {}'.format(row_num))
        # Cost centre column
        if not row[1]:
            failures.append('Cost centre is blank for row {}'.format(row_num))
            logger.warn('Cost centre is blank for row {}'.format(row_num))
        else:
            # Account for text-encodings (Latin1), trailing spaces, etc.
            cc = unidecode(str(row[1])).strip()
            try:
                cc = int(cc)
            except ValueError:  # CC can't be cast as integer.
                failures.append('Cost centre {} is invalid for row {}'.format(
                    cc, row_num))
                logger.warn('Cost centre {} is invalid for row {}'.format(
                    cc, row_num))

        # Fatal validation errors: log any failures and skip to the next row.
        if failures:
            for f in failures:
                LprValidation.objects.create(lpr=lpr, row=row_num, validation=f)
            continue

        # No validation errors so far; try matching a Computer or Mobile.
        computer = None
        mobile = None

        # Computer
        qs = Computer.objects.filter(serial_number=serial)
        if qs.count() > 1:
            failures.append('Serial no {} for row {} matches >1 computer in the db'.format(
                serial, row_num))
            logger.warn('Serial no {} for row {} matches >1 computer in the db'.format(
                serial, row_num))
        elif qs.exists():
            computer = Computer.objects.get(serial_number=serial)
            cc_no = computer.cost_centre_no
            if cc_no and cc_no != cc:  # Already assigned to a CC, <> the CC supplied.
                failures.append('''Computer with serial {} for row {} is
                        assigned to cost centre {}; cannot reassign to {}'''.format(
                    serial, row_num, cc_no, cc))
                logger.warn('''Computer with serial {} for row {} is assigned
                        to cost centre {}; cannot reassign to {}'''.format(
                    serial, row_num, cc_no, cc))
            # If CC is not set, do so now.
            elif not cc_no:  # CC not set
                computer.cost_centre_no = cc
                computer.save()
                logger.info('Computer {} cost centre set to {}.'.format(computer, cc))
        # Mobile
        qs = Mobile.objects.filter(serial_number=serial)
        if qs.count() > 1:
            failures.append('Serial no {} for row {} matches >1 mobile in the db'.format(
                serial, row_num))
            logger.warn('Serial no {} for row {} matches >1 mobile in the db'.format(
                serial, row_num))
        elif qs.exists():
            mobile = Mobile.objects.get(serial_number=serial)
            cc_no = mobile.cost_centre_no
            if cc_no and cc_no != cc:
                failures.append('''Mobile with serial {} for row {} is assigned
                        to cost centre {}; cannot reassign to {}'''.format(
                    serial, row_num, cc_no, cc))
                logger.warn('''Mobile with serial {} for row {} is assigned to
                        cost centre {}; cannot reassign to {}'''.format(
                    serial, row_num, cc_no, cc))
            # If CC is not set, do so now.
            elif not cc_no:  # CC not set
                mobile.cost_centre_no = cc
                mobile.save()
                logger.info('Mobile {} cost centre set to {}.'.format(mobile, cc))

        # Serial number couldn't be matched to any object.
        if not computer and not mobile:
            failures.append('''Serial no {} for row {} does not match any
                single computer or mobile in the db'''.format(serial, row_num))
            logger.warn('''Serial no {} for row {} does not match any single
                computer or mobile in the db'''.format(serial, row_num))

        for f in failures:
            LprValidation.objects.create(lpr=lpr, row=row_num, validation=f)
        """
        # Check uploaded OIM asset no. value.
        if serial and computer and sheet.cell(i, 3).value:
            new_val = str(sheet.cell(i, 3).value)
            if len(new_val) > 64:  # Asset no. too long.
                err_messages.append('''OIM asset number {} for serial no. {}
                    is invalid - please correct the data.'''.format(
                    new_val, serial))
                new_val = None
            if new_val and computer.asset_id and new_val != computer.asset_id:
                old_val = computer.asset_id
                computer.asset_id = new_val
                logger.info('''OIM asset number updated from {} to
                    {}.'''.format(old_val, new_val))
            elif new_val and not computer.asset_id:  # Asset no not set.
                computer.asset_id = new_val
                logger.info('''OIM asset number set to {}.'''.format(
                    new_val))

        # Check uploaded finance asset no. value.
        if serial and computer and sheet.cell(i, 4).value:
            new_val = sheet.cell(i, 4).value
            # Handle Excel weirdness: 'integer' cell values will be floats
            # here. If so, turn the value into an integer, then a string.
            if isinstance(new_val, float):
                new_val = str(int(new_val))
            if len(new_val) > 64:  # Finance asset no. too long.
                err_messages.append('''Finance asset number {} for serial no.
                    {} is invalid - please correct the data.'''.format(
                    new_val, serial))
                new_val = None
            if new_val and computer.finance_asset_id and new_val != computer.finance_asset_id:
                old_val = computer.finance_asset_id
                computer.finance_asset_id = new_val
                logger.info('''Finance asset number updated from {} to
                    {}.'''.format(old_val, new_val))
            elif new_val and not computer.finance_asset_id:  # Asset no not set.
                computer.finance_asset_id = new_val
                logger.info('''Finance asset number set to {}.'''.format(
                    new_val))

        if serial and computer:
            if err_messages:
                computer.validation_notes = ' '.join(err_messages)
            else:
                computer.validation_notes = 'No errors'
            computer.save()
        """
        # TODO: email the uploading user a list of any validation errors.
    return True
