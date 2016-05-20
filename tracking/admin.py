from datetime import datetime
from django import forms
from django.conf.urls import url
from django.contrib import admin, messages
from django.shortcuts import redirect
from django.template.response import TemplateResponse
from openpyxl import load_workbook
import tempfile

from tracking.models import DepartmentUser, Computer, Mobile, EC2Instance
from tracking.utils import logger_setup


class DepartmentUserAdmin(admin.ModelAdmin):
    list_display = ['email', 'employee_id', 'username', 'active', 'vip', 'executive',
                    'cost_centre', 'date_updated', 'date_ad_updated', 'org_unit', 'parent']
    list_filter = ['active', 'vip', 'executive', 'date_ad_updated']
    search_fields = ['name', 'email', 'username', 'employee_id']
    raw_id_fields = ['parent', 'cost_centre', 'org_unit']
    readonly_fields = [
        'username', 'email', 'org_data_pretty', 'ad_data_pretty',
        'active', 'in_sync', 'ad_deleted', 'date_ad_updated', 'expiry_date',
        'alesco_data']
    fieldsets = (
        ('Email/username', {
            'fields': ('email', 'username'),
        }),
        ('Name fields', {
            'description': '''<p class="errornote">Do not edit information in this section
            without written permission from People Services or the cost centre manager
            (forms are required).</p>''',
            'fields': (
                ('given_name', 'surname'),
                ('employee_id', 'cost_centre'),
                ('name', 'org_unit'),
                ('name_update_reference'),
            ),
        }),
        ('Other details', {
            'fields': (
                'photo',
                'account_type',
                ('cost_centres_secondary', 'org_units_secondary'),
                ('telephone', 'mobile_phone', 'other_phone'),
                ('populate_primary_group', 'vip', 'executive', 'contractor'),
                ('title', 'parent'),
                'secondary_locations',
                'notes',
                'working_hours',
                'extra_data',
            )
        }),
        ('Organisation data (read-only)', {
            'fields': (
                ('active', 'in_sync', 'ad_deleted', 'date_ad_updated', 'expiry_date'),
                'org_data_pretty',
                'ad_data_pretty',
                'alesco_data',
            )
        })
    )

    def save_model(self, request, obj, form, change):
        """Override save_model in order to log any changes to some fields:
        'given_name', 'surname', 'employee_id', 'cost_centre', 'name', 'org_unit'
        """
        logger = logger_setup('departmentuser_updates')
        l = 'DepartmentUser: {}, field: {}, original_value: {} new_value: {}, changed_by: {}, reference: {}'
        if obj._DepartmentUser__original_given_name != obj.given_name:
            logger.info(l.format(
                obj.email, 'given_name', obj._DepartmentUser__original_given_name, obj.given_name,
                request.user.username, obj.name_update_reference
            ))
        if obj._DepartmentUser__original_surname != obj.surname:
            logger.info(l.format(
                obj.email, 'surname', obj._DepartmentUser__original_surname, obj.surname,
                request.user.username, obj.name_update_reference
            ))
        if obj._DepartmentUser__original_employee_id != obj.employee_id:
            logger.info(l.format(
                obj.email, 'employee_id', obj._DepartmentUser__original_employee_id,
                obj.employee_id, request.user.username, obj.name_update_reference
            ))
        if obj._DepartmentUser__original_cost_centre != obj.cost_centre:
            logger.info(l.format(
                obj.email, 'cost_centre', obj._DepartmentUser__original_cost_centre,
                obj.cost_centre, request.user.username, obj.name_update_reference
            ))
        if obj._DepartmentUser__original_name != obj.name:
            logger.info(l.format(
                obj.email, 'name', obj._DepartmentUser__original_name, obj.name,
                request.user.username, obj.name_update_reference
            ))
        if obj._DepartmentUser__original_org_unit != obj.org_unit:
            logger.info(l.format(
                obj.email, 'org_unit', obj._DepartmentUser__original_org_unit, obj.org_unit,
                request.user.username, obj.name_update_reference
            ))
        obj.save()

    def get_urls(self):
        urls = super(DepartmentUserAdmin, self).get_urls()
        urls = [
            url(r'^alesco-import/$', self.alesco_import, name='alesco_import'),
        ] + urls
        return urls

    class AlescoImportForm(forms.Form):
        spreadsheet = forms.FileField()

    def alesco_import(self, request):
        """Displays a form prompting user to upload an Excel spreadsheet of
        employee data from Alesco. Temporary measure until database link is
        worked out.
        """
        context = dict(
            admin.site.each_context(request),
            title='Alesco data import'
        )

        if request.method == 'POST':
            form = self.AlescoImportForm(request.POST, request.FILES)
            if form.is_valid():
                upload = request.FILES['spreadsheet']
                # Write the uploaded file to a temp file.
                f = tempfile.NamedTemporaryFile(suffix='.' + upload.name.split('.')[-1])
                f.write(upload.read())
                f.seek(0)
                wb = load_workbook(filename=f.name, read_only=True)
                ws = wb.worksheets[0]
                keys = []
                values = []
                non_matched = 0
                multi_matched = 0
                updates = 0
                # Iterate over each row in the worksheet.
                for k, row in enumerate(ws.iter_rows()):
                    values = []
                    for cell in row:
                        # First row: generate keys.
                        if k == 0:
                            keys.append(cell.value)
                        # Otherwise make a list of values.
                        else:
                            # Serialise datetime objects.
                            if isinstance(cell.value, datetime):
                                values.append(cell.value.isoformat())
                            else:
                                values.append(cell.value)
                    if k > 0:
                        # Construct a dictionary of row values.
                        record = dict(zip(keys, values))
                        # Try to find a matching DepartmentUser by employee id.
                        d = DepartmentUser.objects.filter(employee_id=record['EMPLOYEE_NO'])
                        if d.count() > 1:
                            multi_matched += 1
                        elif d.count() == 1:
                            d = d[0]
                            d.alesco_data = record
                            d.save()
                            updates += 1
                        else:
                            non_matched += 0
                # Messages then redirect to DepartmentUser change_list
                if updates > 0:
                    messages.success(request, 'Alesco data for {} DepartmentUsers was updated.'.format(updates))
                if non_matched > 0:
                    messages.warning(request, 'Employee ID was not matched for {} rows.'.format(non_matched))
                if multi_matched > 0:
                    messages.error(request, 'Employee ID was matched for >1 DepartmentUsers for {} rows.'.format(multi_matched))
                return redirect('admin:tracking_departmentuser_changelist')
        else:
            form = self.AlescoImportForm()
        context['form'] = form

        return TemplateResponse(request, 'tracking/alesco_import.html', context)


class ComputerAdmin(admin.ModelAdmin):
    list_display = ['sam_account_name',
                    'hostname', 'managed_by', 'probable_owner']
    search_fields = ['sam_account_name', 'hostname']


class MobileAdmin(admin.ModelAdmin):
    list_display = ('identity', 'model', 'imei', 'serial_number',
                    'asset_id', 'finance_asset_id', 'registered_to')
    search_fields = ('identity', 'model', 'imei',
                     'serial_number', 'asset_id', 'finance_asset_id')


class EC2InstanceAdmin(admin.ModelAdmin):
    list_display = ('name', 'ec2id', 'launch_time', 'running', 'next_state')
    search_fields = ('name', 'ec2id', 'launch_time')
    readonly_fields = ["extra_data_pretty", "extra_data"]


admin.site.register(DepartmentUser, DepartmentUserAdmin)
admin.site.register(Computer, ComputerAdmin)
admin.site.register(Mobile, MobileAdmin)
admin.site.register(EC2Instance, EC2InstanceAdmin)
