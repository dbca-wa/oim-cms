import csv
from datetime import datetime, timedelta
from django.conf import settings
from django.core.files.base import ContentFile
import logging
from openpyxl import load_workbook
import os
from StringIO import StringIO
import subprocess
import unicodecsv

from tracking.models import DepartmentUser


def load_mugshots(data_dir='/root/mugshots'):
    files = [x for x in os.listdir(data_dir) if os.path.isfile(os.path.join(data_dir, x))]
    valid = 0
    for f in files:
        name = os.path.splitext(f)[0]
        qs = DepartmentUser.objects.filter(username__iexact=name)
        if qs:
            with open(os.path.join(data_dir, f)) as fp:
                qs[0].photo.save(f, ContentFile(fp.read()))
            print('Updated photo for {}'.format(name))
            valid += 1
        else:
            print('ERROR: Username {} not found'.format(name))

    print('{}/{} photos valid'.format(valid, len(files)))


def convert_ad_timestamp(timestamp):
    """Converts an Active Directory timestamp to a Python datetime object.
    Ref: http://timestamp.ooz.ie/p/time-in-python.html
    """
    epoch_start = datetime(year=1601, month=1, day=1)
    seconds_since_epoch = timestamp / 10**7
    return epoch_start + timedelta(seconds=seconds_since_epoch)


def logger_setup(name):
    # Ensure that the logs dir is present.
    subprocess.check_call(['mkdir', '-p', 'logs'])
    # Set up logging in a standardised way.
    logger = logging.getLogger(name)
    if settings.DEBUG:
        logger.setLevel(logging.DEBUG)
    else:  # Log at a higher level when not in debug mode.
        logger.setLevel(logging.INFO)
    if not len(logger.handlers):  # Avoid creating duplicate handlers.
        fh = logging.handlers.RotatingFileHandler(
            'logs/{}.log'.format(name), maxBytes=5 * 1024 * 1024, backupCount=5)
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        fh.setFormatter(formatter)
        logger.addHandler(fh)
    return logger


def csv_data(csv_path, skip_header=True):
    """Pass in the path to a CSV file, returns a CSV Reader object.
    """
    csv_file = open(csv_path, 'r')
    # Determine the CSV dialect.
    dialect = unicodecsv.Sniffer().sniff(csv_file.read(1024))
    csv_file.seek(0)
    data = unicodecsv.reader(csv_file, dialect)
    if skip_header:
        data.next()
    return data


def sourcefile_setup(sourcefile):
    """Read source CSV data (from source_files dir), return a csvreader object
    and a count of the rows. Assumes that data has a header.
    """
    source = os.path.join(settings.BASE_DIR, 'source_files', sourcefile)
    csvfile = open(source, 'r')
    csvcounter = csv.reader(csvfile)
    row_count = sum(1 for row in csvcounter) - 1  # Read the file to count rows
    csvfile.close()
    # Reopen the CSV.
    csvfile = open(source, 'r')
    csvreader = csv.reader(csvfile)
    csvreader.next()  # Skip the headers row.
    return csvreader, row_count


def strip_non_ascii(string):
    """Returns a string without non-ASCII characters.
    """
    stripped = (c for c in string if 0 < ord(c) < 127)
    return ''.join(stripped)


def csv_sync_prop_register_data(src='it_assets.csv'):
    """Utility function to import local property register data from a CSV file
    (called it_assets.csv by default) saved in the source_files dir.

    Matches on serial number, should be non-destructive of existing data.
    """
    from .models import Computer
    data, count = sourcefile_setup(src)

    for i in data:
        serial = i[0].strip().upper()  # We save uppercase serials in the db.
        if serial:  # Non-blanks
            comps = Computer.objects.filter(serial_number=serial)
        else:
            comps = None
        if comps:
            # Cope with having duplicate serial numbers in Incredibus.
            for c in comps:
                if i[1]:  # Cost centre no.
                    c.cost_centre_no = int(i[1])
                if i[3]:  # Finance asset no.
                    c.finance_asset_id = i[3]
                if i[4]:  # OIM asset no.
                    c.asset_id = i[4]
                c.save()
                print('Updated {}'.format(c))
                if c.cost_centre_no:
                    c.ad_set_cost_centre()
                    print('Updated {} CC in AD'.format(c))
                if c.finance_asset_id:
                    c.ad_set_finance_asset_number()
                    print('Updated {} Finance asset no in AD'.format(c))
                if c.asset_id:
                    c.ad_set_asset_number()
                    print('Updated {} OIM asset no in AD'.format(c))

    print('Done')


def alesco_data_import(filepath):
    """Import task expects to be passed a file path to a closed .xlsx file.
    """
    logger = logger_setup('alesco_data_import')
    f = open(filepath)
    wb = load_workbook(filename=f.name, read_only=True)
    ws = wb.worksheets[0]
    keys = []
    values = []
    non_matched = 0
    multi_matched = 0
    updates = 0
    # Iterate over each row in the worksheet.
    for k, row in enumerate(ws.iter_rows()):
        values = []
        for cell in row:
            # First row: generate keys.
            if k == 0:
                keys.append(cell.value)
            # Otherwise make a list of values.
            else:
                # Serialise datetime objects.
                if isinstance(cell.value, datetime):
                    values.append(cell.value.isoformat())
                else:
                    values.append(cell.value)
        if k > 0:
            # Construct a dictionary of row values.
            record = dict(zip(keys, values))
            # Try to find a matching DepartmentUser by employee id.
            d = DepartmentUser.objects.filter(employee_id=record['EMPLOYEE_NO'])
            if d.count() > 1:
                multi_matched += 1
            elif d.count() == 1:
                d = d[0]
                d.alesco_data = record
                d.save()
                logger.info('Alesco data updated for {}'.format(d.email.lower()))
                updates += 1
            else:
                non_matched += 0
    if updates > 0:
        logger.info('Alesco data for {} DepartmentUsers was updated.'.format(updates))
    if non_matched > 0:
        logger.warning('Employee ID was not matched for {} rows.'.format(non_matched))
    if multi_matched > 0:
        logger.error('Employee ID was matched for >1 DepartmentUsers for {} rows.'.format(multi_matched))

    return True


def departmentuser_csv_report():
    """Output data from all DepartmentUser objects to a CSV, unpacking the
    various JSONField values.
    Returns a StringIO object that can be written to a response or file.
    """
    FIELDS = [
        'email', 'username', 'given_name', 'surname', 'name', 'employee_id',
        'cost_centre', 'org_unit', 'name_update_reference',
        'telephone', 'mobile_phone', 'other_phone', 'title', 'preferred_name',
        'security_clearance']
    TYPE_CHOICES = {x[0]: x[1] for x in DepartmentUser.ACCOUNT_TYPE_CHOICES}

    # Get any DepartmentUser with non-null alesco_data field.
    # alesco_data structure should be consistent to all (or null).
    du = DepartmentUser.objects.filter(alesco_data__isnull=False)[0]
    alesco_fields = du.alesco_data.keys()
    alesco_fields.sort()
    org_fields = {
        'department': ('units', 0, 'name'),
        'division': ('units', 1, 'name'),
        'branch': ('units', 2, 'name')
    }

    header = [f for f in FIELDS]
    # These fields appended manually:
    header.append('account_type')
    header.append('position_type')
    header.append('reports_to')
    header += org_fields.keys()
    header += alesco_fields

    # Get a DepartmentUser with non-null org_data field, for keys.
    du = DepartmentUser.objects.filter(org_data__isnull=False)[0]
    cc_keys = du.org_data['cost_centre'].keys()
    header += ['cost_centre_{}'.format(k) for k in cc_keys]
    location_keys = du.org_data['location'].keys()
    header += ['location_{}'.format(k) for k in location_keys]
    header.append('secondary_location')

    # Get a DepartmentUser with non-null ad_data field, for keys.
    du = DepartmentUser.objects.filter(ad_data__isnull=False)[0]
    ad_keys = du.ad_data.keys()
    ad_keys.remove('mailbox')  # Remove the nested object.
    header += ['ad_{}'.format(k) for k in ad_keys]

    # Write data for all DepartmentUser objects to the CSV
    stream = StringIO()
    wr = unicodecsv.writer(stream, encoding='utf-8')
    wr.writerow(header)
    for u in DepartmentUser.objects.filter(active=True):
        record = []
        for f in FIELDS:
            record.append(getattr(u, f))
        try:  # Append account_type display value.
            record.append(TYPE_CHOICES[u.account_type])
        except:
            record.append('')
        try:  # Append position_type display value.
            record.append(TYPE_CHOICES[u.position_type])
        except:
            record.append('')
        record.append(u.parent)  # Append parent field.
        for o in org_fields:
            try:
                src = u.org_data
                for x in org_fields[o]:
                    src = src[x]
                record.append(src)
            except:
                record.append('')

        for a in alesco_fields:
            try:
                record.append(u.alesco_data[a])
            except:
                record.append('')
        for i in cc_keys:
            try:
                record.append(u.org_data['cost_centre'][i])
            except:
                record.append('')
        for i in location_keys:
            try:
                record.append(u.org_data['location'][i])
            except:
                record.append('')
        if u.org_data and 'secondary_location' in u.org_data:
            record.append(u.org_data['secondary_location'])
        else:
            record.append('')
        for i in ad_keys:
            try:
                record.append(u.ad_data[i])
            except:
                record.append('')

        # Write the row to the CSV stream.
        wr.writerow(record)

    return stream.getvalue()
