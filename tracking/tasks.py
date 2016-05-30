from __future__ import absolute_import
from celery import shared_task
from datetime import datetime
from openpyxl import load_workbook

from tracking.models import DepartmentUser
from tracking.utils import logger_setup


logger = logger_setup('alesco_data_import')


@shared_task(name='alesco-data-import')
def alesco_data_import(filepath):
    """Import task expects to be passed a file path to a closed .xlsx file.
    """
    f = open(filepath)
    wb = load_workbook(filename=f.name, read_only=True)
    ws = wb.worksheets[0]
    keys = []
    values = []
    non_matched = 0
    multi_matched = 0
    updates = 0
    # Iterate over each row in the worksheet.
    for k, row in enumerate(ws.iter_rows()):
        values = []
        for cell in row:
            # First row: generate keys.
            if k == 0:
                keys.append(cell.value)
            # Otherwise make a list of values.
            else:
                # Serialise datetime objects.
                if isinstance(cell.value, datetime):
                    values.append(cell.value.isoformat())
                else:
                    values.append(cell.value)
        if k > 0:
            # Construct a dictionary of row values.
            record = dict(zip(keys, values))
            # Try to find a matching DepartmentUser by employee id.
            d = DepartmentUser.objects.filter(employee_id=record['EMPLOYEE_NO'])
            if d.count() > 1:
                multi_matched += 1
            elif d.count() == 1:
                d = d[0]
                d.alesco_data = record
                d.save()
                updates += 1
            else:
                non_matched += 0
    if updates > 0:
        logger.info('Alesco data for {} DepartmentUsers was updated.'.format(updates))
    if non_matched > 0:
        logger.warning('Employee ID was not matched for {} rows.'.format(non_matched))
    if multi_matched > 0:
        logger.error('Employee ID was matched for >1 DepartmentUsers for {} rows.'.format(multi_matched))

    return True
